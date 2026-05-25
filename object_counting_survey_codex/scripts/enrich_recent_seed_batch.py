"""Enrich recently added Scholar seeds with DOI, notes, BibTeX, and PDFs."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SEED_FILE = ROOT / "data" / "seed_papers_google_scholar.xlsx"
BIBTEX_JSON = Path("/private/tmp/recent_seed_bibtex_browser.json")

FOLDER_BY_SHEET = {
    "Videos": "video",
    "Others": "other_modalities",
    "Benchmarks_Datasets": "benchmarks_datasets",
    "Remote_Sensing_Aerial_UAV": "remote_sensing_aerial_uav",
    "Medical_Microscopy_Cell": "medical_microscopy_cell",
    "Thermal_Event_Camera": "thermal_event_camera",
}

MANUAL_DOI = {
    "Video object counting with scene-aware multi-object tracking": "10.4018/IJDST.321553",
    "Stacked objects count based on density map and two vanishing points": "10.1117/12.2640364",
    "Rapid characterization of cell and bacteria counts using computer vision": "10.46810/tdfd.902441",
    "TrueCount: Improving Open-World Object Counting with Visual-Language Models and Dynamic Multi-Modal Inputs": "10.1145/3746027.3755426",
}


def safe_filename(seed_id: str, title: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9]+", "_", title).strip("_")[:90]
    return f"{seed_id}_{stem}.pdf"


def title_norm(title: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", title.lower()).strip()


def extract_doi(*values: str) -> str:
    text = " ".join(v or "" for v in values)
    match = re.search(r"(10\.\d{4,9}/[^\s\"<>]+)", text, re.I)
    if not match:
        return ""
    return match.group(1).rstrip(").,;}]")


def doi_from_crossref(title: str) -> str:
    url = "https://api.crossref.org/works"
    try:
        proc = subprocess.run(
            [
                "curl",
                "-L",
                "--fail",
                "--silent",
                "--show-error",
                "--max-time",
                "20",
                "--get",
                "--data-urlencode",
                f"query.title={title}",
                "--data-urlencode",
                "rows=3",
                url,
            ],
            check=False,
            capture_output=True,
            text=True,
        )
        if proc.returncode != 0:
            return ""
        payload = json.loads(proc.stdout)
    except Exception:
        return ""

    wanted = title_norm(title)
    best_doi = ""
    best_score = 0
    for item in payload.get("message", {}).get("items", []):
        candidate = title_norm(" ".join(item.get("title") or []))
        if not candidate:
            continue
        wanted_words = set(wanted.split())
        candidate_words = set(candidate.split())
        if not wanted_words:
            continue
        score = len(wanted_words & candidate_words) / len(wanted_words)
        if score > best_score:
            best_score = score
            best_doi = item.get("DOI", "")
    return best_doi if best_score >= 0.72 else ""


def download_pdf(url: str, dest: Path) -> str:
    if not url:
        return "No PDF URL available."
    if str(url).startswith("papers_pdf/"):
        local = ROOT / str(url)
        if local.exists():
            return f"Saved PDF: {local.relative_to(ROOT)}"
        return "Local PDF path was recorded but file is missing."
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_suffix(".tmp")
    proc = subprocess.run(
        [
            "curl",
            "-L",
            "--fail",
            "--silent",
            "--show-error",
            "--max-time",
            "60",
            "-A",
            "Mozilla/5.0",
            "-o",
            str(tmp),
            url,
        ],
        check=False,
        capture_output=True,
        text=True,
    )
    if proc.returncode != 0:
        if tmp.exists():
            tmp.unlink()
        return f"PDF download failed or requires access: {proc.stderr.strip()[:160]}"
    try:
        header = tmp.read_bytes()[:5]
        if header != b"%PDF-":
            tmp.unlink()
            return "Downloaded response was not a PDF; likely landing page, paywall, or login page."
        shutil.move(str(tmp), dest)
        return f"Saved PDF: {dest.relative_to(ROOT)}"
    except Exception as exc:
        if tmp.exists():
            tmp.unlink()
        return f"PDF verification failed: {exc}"


def main() -> None:
    bib_records = {}
    if BIBTEX_JSON.exists():
        for item in json.loads(BIBTEX_JSON.read_text()):
            bib_records[item["seed_id"]] = item

    wb = load_workbook(SEED_FILE)
    updated = 0
    pdf_saved = 0
    pdf_issues = []
    doi_count = 0
    bib_ok = 0
    bib_blocked = []

    for sheet_name, folder in FOLDER_BY_SHEET.items():
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        col = {h: i + 1 for i, h in enumerate(headers)}
        for row in range(2, ws.max_row + 1):
            note = str(ws.cell(row, col["notes"]).value or "")
            seed_id = ws.cell(row, col["seed_id"]).value
            if "Added from visible Google Scholar results" not in note and seed_id not in bib_records:
                continue
            title = ws.cell(row, col["paper_title"]).value
            contribution = ws.cell(row, col["main_contribution"]).value or ""
            paper_url = ws.cell(row, col["paper_url"]).value or ""
            pdf_url = ws.cell(row, col["pdf_url"]).value or ""
            if not str(pdf_url).strip() and re.search(r"(\.pdf($|[?])|/content/pdf/|/pdf/)", str(paper_url), re.I):
                pdf_url = paper_url

            doi = (
                MANUAL_DOI.get(title)
                or extract_doi(paper_url, pdf_url, ws.cell(row, col["bibtex"]).value or "")
                or doi_from_crossref(title)
            )
            if doi:
                doi_count += 1
                ws.cell(row, col["doi"], f"https://doi.org/{doi.lower() if doi.startswith('10.') else doi}")

            bib = bib_records.get(seed_id, {})
            if bib.get("status") == "ok" and bib.get("bibtex"):
                bib_ok += 1
                ws.cell(row, col["bibtex"], bib["bibtex"])
                ws.cell(row, col["bibtex_collected"], "Yes")
                bib_doi = extract_doi(bib["bibtex"])
                if bib_doi and not ws.cell(row, col["doi"]).value:
                    ws.cell(row, col["doi"], f"https://doi.org/{bib_doi}")
            else:
                ws.cell(row, col["bibtex_collected"], f"No - Scholar {bib.get('status', 'not collected')}")
                bib_blocked.append(seed_id)

            existing_pdf_status = str(ws.cell(row, col["pdf_saved"]).value or "")
            if existing_pdf_status == "Yes" and str(ws.cell(row, col["pdf_url"]).value or "").startswith("papers_pdf/"):
                pdf_status = f"Saved PDF: {ws.cell(row, col['pdf_url']).value}"
            else:
                pdf_folder = ROOT / "papers_pdf" / folder
                pdf_path = pdf_folder / safe_filename(seed_id, title)
                pdf_status = download_pdf(pdf_url, pdf_path)
            if pdf_status.startswith("Saved PDF"):
                pdf_saved += 1
                ws.cell(row, col["pdf_saved"], "Yes")
                ws.cell(row, col["pdf_url"], str(pdf_path.relative_to(ROOT)))
            else:
                ws.cell(row, col["pdf_saved"], f"No - {pdf_status}")
                pdf_issues.append((seed_id, title, pdf_status))

            ws.cell(row, col["notes"], f"Key contribution: {contribution} PDF status: {pdf_status}")
            updated += 1

    wb.save(SEED_FILE)
    print(f"Updated {updated} recently added rows.")
    print(f"DOIs/DOI links filled: {doi_count}")
    print(f"BibTeX collected from Scholar: {bib_ok}")
    print(f"PDFs saved: {pdf_saved}")
    if bib_blocked:
        print("BibTeX still needing manual Scholar retry:", ", ".join(bib_blocked))
    if pdf_issues:
        print("PDF issues:")
        for seed_id, title, issue in pdf_issues:
            print(f"- {seed_id}: {title} -- {issue}")


if __name__ == "__main__":
    main()
