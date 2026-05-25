"""Merge Scholar, Scopus, and API raw records; deduplicate and score."""

from __future__ import annotations

import csv
import re
from collections import Counter
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
RAW_DIR = DATA / "raw_api_results"
SEED_XLSX = DATA / "seed_papers_google_scholar.xlsx"
SCOPUS_IMPORTED = RAW_DIR / "scopus_imported_records.csv"
MASTER_CSV = DATA / "master_papers_raw.csv"
CLEAN_XLSX = DATA / "literature_matrix_clean.xlsx"
SCREENING_XLSX = DATA / "screening_log.xlsx"
CORE_XLSX = ROOT / "outputs" / "core_reading_set.xlsx"
DEDUP_UNIQUE_CSV = RAW_DIR / "deduplicated_unique_records.csv"
DUPLICATES_CSV = RAW_DIR / "duplicate_removed_records.csv"
DEDUP_CHECKPOINT_CSV = RAW_DIR / "dedup_checkpoint_latest.csv"

SCHEMA = [
    "record_id",
    "source",
    "source_query",
    "export_filename",
    "paper_title",
    "authors",
    "year",
    "venue",
    "doi",
    "arxiv_id",
    "paper_url",
    "abstract",
    "author_keywords",
    "index_keywords",
    "citation_count",
    "document_type",
    "dataset_or_benchmark",
    "modality",
    "task_category",
    "method_family",
    "input_type",
    "output_type",
    "application_area",
    "relevance_score",
    "tier",
    "reason_kept_or_removed",
    "bibtex",
    "pdf_saved",
    "github_url",
]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def norm_title(value: object) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def norm_doi(value: object) -> str:
    text = clean_text(value)
    text = re.sub(r"^https?://(dx\.)?doi\.org/", "", text, flags=re.I)
    return text.rstrip("}. ").lower()


def norm_arxiv(value: object) -> str:
    text = clean_text(value)
    text = re.sub(r"^arxiv:", "", text, flags=re.I)
    match = re.search(r"(\d{4}\.\d{4,5})(v\d+)?", text)
    return match.group(1) if match else text.lower()


def as_int(value: object) -> int:
    try:
        return int(float(clean_text(value).replace(",", "")))
    except Exception:
        return 0


def load_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return [dict(row) for row in csv.DictReader(f)]


def load_scholar_seeds() -> list[dict[str, str]]:
    if not SEED_XLSX.exists():
        return []
    wb = load_workbook(SEED_XLSX, read_only=True, data_only=True)
    records = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows or not any(rows[0] or []):
            continue
        header = [clean_text(h) for h in rows[0]]
        if "paper_title" not in header:
            continue
        for idx, row in enumerate(rows[1:], start=1):
            if not any(v not in (None, "") for v in row):
                continue
            values = {header[i]: row[i] if i < len(row) else "" for i in range(len(header))}
            title = clean_text(values.get("paper_title"))
            if not title:
                continue
            records.append(
                {
                    "record_id": clean_text(values.get("seed_id")) or f"SCHOLAR_{ws.title}_{idx:04d}",
                    "source": "Google Scholar",
                    "source_query": clean_text(values.get("google_scholar_query")),
                    "export_filename": "",
                    "paper_title": title,
                    "authors": clean_text(values.get("authors")),
                    "year": clean_text(values.get("year")),
                    "venue": clean_text(values.get("venue")),
                    "doi": norm_doi(values.get("doi")),
                    "arxiv_id": norm_arxiv(values.get("arxiv_id")),
                    "paper_url": clean_text(values.get("paper_url")),
                    "abstract": "",
                    "author_keywords": "",
                    "index_keywords": "",
                    "citation_count": clean_text(values.get("citation_count")),
                    "document_type": "",
                    "dataset_or_benchmark": clean_text(values.get("dataset_or_benchmark")),
                    "modality": clean_text(values.get("modality")),
                    "task_category": clean_text(values.get("task_category")),
                    "method_family": clean_text(values.get("task_category")),
                    "input_type": clean_text(values.get("input_type")),
                    "output_type": clean_text(values.get("output_type")),
                    "application_area": "",
                    "relevance_score": seed_score(values.get("relevance_score")),
                    "tier": "",
                    "reason_kept_or_removed": clean_text(values.get("notes")),
                    "bibtex": clean_text(values.get("bibtex")),
                    "pdf_saved": clean_text(values.get("pdf_saved")),
                    "github_url": clean_text(values.get("github_url")),
                    "seed_bucket": ws.title,
                }
            )
    return records


def seed_score(value: object) -> str:
    text = clean_text(value).lower()
    if text in {"high", "5"}:
        return "5"
    if text in {"medium", "med", "3"}:
        return "3"
    if text in {"low", "1"}:
        return "1"
    return clean_text(value)


def collect_raw_records() -> list[dict[str, str]]:
    records = []
    records.extend(load_scholar_seeds())
    records.extend(load_csv(SCOPUS_IMPORTED))
    for path in sorted(RAW_DIR.glob("*.csv")):
        if path.name == SCOPUS_IMPORTED.name:
            continue
        for row in load_csv(path):
            if not clean_text(row.get("paper_title")):
                continue
            row.setdefault("source", path.stem)
            row.setdefault("record_id", f"{path.stem}_{len(records)+1:06d}")
            records.append(row)
    normalized = []
    for i, row in enumerate(records, start=1):
        item = {key: clean_text(row.get(key, "")) for key in SCHEMA}
        item["record_id"] = item["record_id"] or f"RAW_{i:06d}"
        item["doi"] = norm_doi(item["doi"])
        item["arxiv_id"] = norm_arxiv(item["arxiv_id"])
        item["_normalized_title"] = norm_title(item["paper_title"])
        normalized.append(item)
    return normalized


def infer_missing_taxonomy(row: dict[str, str]) -> None:
    text = " ".join(
        [
            row.get("paper_title", ""),
            row.get("abstract", ""),
            row.get("author_keywords", ""),
            row.get("index_keywords", ""),
            row.get("source_query", ""),
        ]
    ).lower()
    if not row.get("modality"):
        if any(k in text for k in ["cell counting", "microscopy", "bacteria", "histology", "medical image"]):
            row["modality"] = "medical; microscopy"
        elif any(k in text for k in ["remote sensing", "aerial", "uav", "drone", "satellite"]):
            row["modality"] = "remote_sensing; aerial"
        elif any(k in text for k in ["point cloud", "3d object", "rgb-d", "lidar", "depth"]):
            row["modality"] = "3d; point_cloud"
        elif any(k in text for k in ["video", "tracking", "tracklet", "temporal"]):
            row["modality"] = "video"
        elif any(k in text for k in ["thermal", "infrared"]):
            row["modality"] = "thermal"
        elif any(k in text for k in ["event camera", "event-based", "neuromorphic"]):
            row["modality"] = "event_camera"
        elif any(k in text for k in ["vision-language", "open-vocabulary", "text-guided", "zero-shot", "clip", "sam"]):
            row["modality"] = "image; multimodal"
        else:
            row["modality"] = "image"
    if not row.get("task_category") or row["task_category"] in {"API expansion", "DOI/metadata expansion", "Recent/gap-fill preprint"}:
        labels = []
        if any(k in text for k in ["survey", "review"]):
            labels.append("Survey")
        if any(k in text for k in ["benchmark", "dataset", "fsc-147", "fscd-147", "countbench", "omnicount", "carpk", "shanghaitech", "tao-count", "videocount"]):
            labels.append("Dataset/Benchmark")
        if any(k in text for k in ["few-shot", "few shot", "exemplar", "class-agnostic", "class agnostic"]):
            labels.append("Few-shot/Class-agnostic")
        if any(k in text for k in ["zero-shot", "zero shot", "open-vocabulary", "open vocabulary", "text-guided"]):
            labels.append("Zero-shot/Open-vocabulary")
        if any(k in text for k in ["foundation model", "vision-language", "clip", "sam"]):
            labels.append("Foundation/Vision-language")
        if any(k in text for k in ["density map", "density estimation", "crowd counting"]):
            labels.append("Density-based")
        if any(k in text for k in ["tracking", "re-identification", "unique instance"]):
            labels.append("Tracking/Unique-count")
        row["task_category"] = "; ".join(dict.fromkeys(labels)) or row.get("task_category") or "Application-specific"
        row["method_family"] = row["task_category"]
    if not row.get("input_type"):
        if "video" in row["modality"]:
            row["input_type"] = "Video"
        elif "point_cloud" in row["modality"]:
            row["input_type"] = "Point Cloud"
        elif "remote_sensing" in row["modality"]:
            row["input_type"] = "Remote Sensing Image"
        elif "medical" in row["modality"] or "microscopy" in row["modality"]:
            row["input_type"] = "Medical/Microscopy Image"
        else:
            row["input_type"] = "Image"
        if any(k in text for k in ["prompt", "text-guided", "open-vocabulary"]):
            row["input_type"] += "; Text Prompt"
        if any(k in text for k in ["exemplar", "few-shot", "reference"]):
            row["input_type"] += "; Exemplar Boxes/Patches"
    if not row.get("output_type"):
        if any(k in text for k in ["density map", "density estimation"]):
            row["output_type"] = "Density Map + Count"
        elif any(k in text for k in ["track", "trajectory", "unique instance"]):
            row["output_type"] = "Tracks + Unique Count"
        elif any(k in text for k in ["detect", "localization", "bounding box"]):
            row["output_type"] = "Boxes + Count"
        else:
            row["output_type"] = "Count Only"


def choose_better(existing: dict[str, str], candidate: dict[str, str]) -> dict[str, str]:
    score_existing = completeness(existing)
    score_candidate = completeness(candidate)
    if score_candidate > score_existing:
        winner, loser = candidate, existing
    else:
        winner, loser = existing, candidate
    source_parts = []
    for source in [winner.get("source"), loser.get("source")]:
        source_parts.extend(part.strip() for part in (source or "").split(";") if part.strip())
    sources = sorted(set(source_parts))
    winner["source"] = "; ".join(sources)
    for key in SCHEMA:
        if not winner.get(key) and loser.get(key):
            winner[key] = loser[key]
    return winner


def completeness(row: dict[str, str]) -> int:
    important = ["doi", "arxiv_id", "abstract", "author_keywords", "citation_count", "bibtex", "pdf_saved"]
    return sum(1 for key in SCHEMA if row.get(key)) + sum(2 for key in important if row.get(key))


def deduplicate(records: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    kept: list[dict[str, str]] = []
    removed: list[dict[str, str]] = []
    doi_index: dict[str, int] = {}
    arxiv_index: dict[str, int] = {}
    title_index: dict[str, int] = {}
    title_blocks: dict[str, list[int]] = {}
    total = len(records)
    for processed, record in enumerate(records, start=1):
        reason = ""
        match_idx = None
        if record["doi"] and record["doi"] in doi_index:
            match_idx = doi_index[record["doi"]]
            reason = "duplicate DOI"
        elif record["arxiv_id"] and record["arxiv_id"] in arxiv_index:
            match_idx = arxiv_index[record["arxiv_id"]]
            reason = "duplicate arXiv ID"
        elif record["_normalized_title"] and record["_normalized_title"] in title_index:
            match_idx = title_index[record["_normalized_title"]]
            reason = "duplicate normalized title"
        else:
            candidates = []
            for block_key in blocking_keys(record["_normalized_title"]):
                candidates.extend(title_blocks.get(block_key, []))
            for idx in dict.fromkeys(candidates):
                existing = kept[idx]
                if title_similarity(record["_normalized_title"], existing["_normalized_title"]) >= 0.92:
                    match_idx = idx
                    reason = "fuzzy title duplicate >= 92%"
                    break
        if match_idx is None:
            kept.append(record)
            idx = len(kept) - 1
            if record["doi"]:
                doi_index[record["doi"]] = idx
            if record["arxiv_id"]:
                arxiv_index[record["arxiv_id"]] = idx
            if record["_normalized_title"]:
                title_index[record["_normalized_title"]] = idx
                for block_key in blocking_keys(record["_normalized_title"]):
                    title_blocks.setdefault(block_key, []).append(idx)
        else:
            removed_record = dict(record)
            removed_record["tier"] = "Exclude"
            removed_record["reason_kept_or_removed"] = reason
            removed.append(removed_record)
            kept[match_idx] = choose_better(kept[match_idx], record)
        if processed % 500 == 0 or processed == total:
            print(
                f"Dedup progress: {processed}/{total} records processed; "
                f"{len(kept)} unique; {len(removed)} duplicates removed",
                flush=True,
            )
            write_csv(DEDUP_CHECKPOINT_CSV, kept)
    return kept, removed


def blocking_keys(title: str) -> list[str]:
    if not title:
        return []
    words = [word for word in title.split() if len(word) > 2]
    keys = []
    if len(words) >= 2:
        keys.append(f"first2:{' '.join(words[:2])}")
    elif words:
        keys.append(f"single:{words[0]}:{title[:18]}")
    keys.append(f"prefix:{title[:18]}")
    return keys


def title_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def score_and_tier(row: dict[str, str]) -> None:
    infer_missing_taxonomy(row)
    text = " ".join(
        [
            row.get("paper_title", ""),
            row.get("abstract", ""),
            row.get("author_keywords", ""),
            row.get("index_keywords", ""),
            row.get("task_category", ""),
            row.get("modality", ""),
        ]
    ).lower()
    score = as_int(row.get("relevance_score"))
    if not score:
        score = 0
        if any(k in text for k in ["object counting", "visual counting", "class-agnostic counting", "few-shot counting"]):
            score += 2
        if any(k in text for k in ["counting objects", "object count", "count objects", "cell counting", "crowd counting"]):
            score += 1
        if any(k in text for k in ["fsc-147", "countbench", "omn​​icount", "omnicount", "t ao-count", "tao-count", "videocount"]):
            score += 1
        if any(k in text for k in ["survey", "benchmark", "dataset", "open-vocabulary", "zero-shot", "foundation model", "video object counting"]):
            score += 1
        if any(k in text for k in ["cell counting", "microscopy", "remote sensing", "point cloud", "rgb-d"]):
            score += 1
        if row.get("source") == "Google Scholar":
            score += 1
        if as_int(row.get("citation_count")) >= 50:
            score += 1
        if "object detection" in text and "count" not in text:
            score -= 2
        if any(k in text for k in ["galaxies", "cosmos", "gene expression", "protein", "spectroscopy"]) and "object counting" not in text:
            score -= 1
    score = max(0, min(score, 5))
    row["relevance_score"] = str(score)
    if score >= 5:
        row["tier"] = "A_core"
    elif score >= 4:
        row["tier"] = "B_important"
    elif score >= 2:
        row["tier"] = "C_background"
    else:
        row["tier"] = "Exclude"
    if not row.get("reason_kept_or_removed"):
        row["reason_kept_or_removed"] = f"Auto-scored as {row['tier']} from title/abstract/query metadata."


def ranking_key(row: dict[str, str]) -> tuple[int, int, int]:
    source_bonus = 0
    source = row.get("source", "")
    if "Google Scholar" in source:
        source_bonus += 3
    if "Scopus" in source:
        source_bonus += 2
    if row.get("doi") or row.get("arxiv_id"):
        source_bonus += 1
    return (as_int(row.get("relevance_score")), source_bonus, as_int(row.get("citation_count")))


def write_csv(path: Path, records: list[dict[str, str]]) -> None:
    fields = SCHEMA + ["_normalized_title"]
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)


def write_workbook(path: Path, records: list[dict[str, str]], excluded: list[dict[str, str]] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "included"
    ws.append(SCHEMA)
    for row in records:
        ws.append([row.get(key, "") for key in SCHEMA])
    if excluded is not None:
        ex = wb.create_sheet("excluded_duplicates")
        ex.append(SCHEMA)
        for row in excluded:
            ex.append([row.get(key, "") for key in SCHEMA])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    records = collect_raw_records()
    if not records:
        raise SystemExit("No raw records found. Run 00_import_scopus_exports.py first.")
    write_csv(MASTER_CSV, records)
    kept, removed = deduplicate(records)
    write_csv(DEDUP_UNIQUE_CSV, kept)
    write_csv(DUPLICATES_CSV, removed)
    for row in kept:
        score_and_tier(row)
    initial_excluded = [row for row in kept if row["tier"] == "Exclude"]
    eligible = [row for row in kept if row["tier"] != "Exclude"]
    eligible.sort(key=ranking_key, reverse=True)
    included = eligible[:700]
    overflow = eligible[700:]
    for idx, row in enumerate(included):
        if idx < 150 and as_int(row.get("relevance_score")) >= 4:
            row["tier"] = "A_core"
        elif idx < 350:
            row["tier"] = "B_important"
        else:
            row["tier"] = "C_background"
    for row in overflow:
        row["tier"] = "Exclude"
        row["reason_kept_or_removed"] = "Outside the top 700 automated Day 2 candidates; retained in raw master for later gap-fill/manual rescue."
    excluded = removed + initial_excluded + overflow
    write_workbook(CLEAN_XLSX, included, excluded)
    write_workbook(SCREENING_XLSX, included, excluded)
    core = [row for row in included if row["tier"] == "A_core"]
    write_workbook(CORE_XLSX, core)
    print(f"Raw records written: {len(records)} -> {MASTER_CSV}")
    print(f"Records after deduplication: {len(kept)} -> {DEDUP_UNIQUE_CSV}")
    print(f"Records removed as duplicates: {len(removed)} -> {DUPLICATES_CSV}")
    print(f"Deduplicated included working candidates: {len(included)} -> {CLEAN_XLSX}")
    print(f"Excluded/duplicate/overflow records: {len(excluded)}")
    print(f"Records with DOI after deduplication: {sum(1 for row in kept if row.get('doi'))}")
    print(f"Records missing DOI after deduplication: {sum(1 for row in kept if not row.get('doi'))}")
    print(f"Records with abstract after deduplication: {sum(1 for row in kept if row.get('abstract'))}")
    print(f"Records missing abstract after deduplication: {sum(1 for row in kept if not row.get('abstract'))}")
    print("\nCounts by source:")
    for key, value in Counter(row["source"] for row in included).most_common():
        print(f"  {key or 'UNKNOWN'}: {value}")
    print("\nCounts by modality:")
    for key, value in Counter(row["modality"] for row in included).most_common():
        print(f"  {key or 'UNKNOWN'}: {value}")
    print("\nCounts by tier:")
    for key, value in Counter(row["tier"] for row in included).most_common():
        print(f"  {key or 'UNKNOWN'}: {value}")


if __name__ == "__main__":
    main()
