"""Normalize the manual seed workbook to the final guide schema."""

from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SEED_FILE = ROOT / "data" / "seed_papers_google_scholar.xlsx"

GUIDE_COLUMNS = [
    "seed_id",
    "paper_title",
    "authors",
    "year",
    "venue",
    "source_found",
    "google_scholar_query",
    "scopus_query",
    "modality",
    "task_category",
    "input_type",
    "output_type",
    "dataset_or_benchmark",
    "main_contribution",
    "relevance_score",
    "doi",
    "arxiv_id",
    "paper_url",
    "bibtex_collected",
    "pdf_saved",
    "notes",
    "citation_count",
    "pdf_url",
    "bibtex",
    "github_url",
    "seed_bucket",
]

EXTRA_SHEETS = [
    "Benchmarks_Datasets",
    "Remote_Sensing_Aerial_UAV",
    "Medical_Microscopy_Cell",
    "Thermal_Event_Camera",
]


def norm_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def cell_value(row: list[object], header_map: dict[str, int], *names: str) -> object:
    for name in names:
        idx = header_map.get(norm_header(name))
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def extract_arxiv(*values: object) -> str:
    text = " ".join(str(v or "") for v in values)
    match = re.search(r"arxiv[:./ ]+(\d{4}\.\d{4,5})(?:v\d+)?", text, re.I)
    return match.group(1) if match else ""


def clean_doi(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if "arxiv" in text.lower():
        return ""
    text = text.replace("https://doi.org/", "").replace("http://doi.org/", "")
    return text.strip().rstrip("}")


def yes_no(value: object) -> str:
    text = str(value or "").strip()
    return "Yes" if text and text.upper() not in {"N/A", "NA", "NONE"} else "No"


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = {
        "seed_id": 12,
        "paper_title": 48,
        "authors": 42,
        "main_contribution": 52,
        "notes": 52,
        "google_scholar_query": 34,
        "scopus_query": 34,
        "paper_url": 42,
        "pdf_url": 42,
        "bibtex": 48,
    }
    for idx, cell in enumerate(ws[1], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(cell.value, 22)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def normalize_sheet(source_ws, target_ws, prefix: str) -> int:
    target_ws.append(GUIDE_COLUMNS)
    headers = [cell.value for cell in source_ws[1]]
    header_map = {norm_header(h): idx for idx, h in enumerate(headers) if h}
    count = 0
    for raw in source_ws.iter_rows(min_row=2, values_only=True):
        title = cell_value(raw, header_map, "paper_title", "Title")
        if not title:
            continue
        count += 1
        doi_raw = cell_value(raw, header_map, "doi", "DOI")
        url = cell_value(raw, header_map, "paper_url", "URL")
        pdf_url = cell_value(raw, header_map, "pdf_url")
        bibtex = cell_value(raw, header_map, "bibtex")
        row = {
            "seed_id": f"{prefix}{count:03d}",
            "paper_title": title,
            "authors": cell_value(raw, header_map, "authors", "Authors"),
            "year": cell_value(raw, header_map, "year", "Year"),
            "venue": cell_value(raw, header_map, "venue", "Venue"),
            "source_found": "Google Scholar",
            "google_scholar_query": cell_value(raw, header_map, "google_scholar_query", "query used"),
            "scopus_query": "",
            "modality": cell_value(raw, header_map, "modality", "Modality"),
            "task_category": cell_value(raw, header_map, "task_category", "Task_category"),
            "input_type": cell_value(raw, header_map, "input_type", "Input_type"),
            "output_type": cell_value(raw, header_map, "output_type", "Output_type"),
            "dataset_or_benchmark": cell_value(raw, header_map, "dataset_or_benchmark", "Dataset"),
            "main_contribution": cell_value(raw, header_map, "main_contribution", "Notes"),
            "relevance_score": cell_value(raw, header_map, "relevance_score"),
            "doi": clean_doi(doi_raw),
            "arxiv_id": extract_arxiv(doi_raw, url, pdf_url, bibtex),
            "paper_url": url,
            "bibtex_collected": yes_no(bibtex),
            "pdf_saved": yes_no(pdf_url),
            "notes": cell_value(raw, header_map, "notes", "Notes"),
            "citation_count": cell_value(raw, header_map, "citation_count", "Citation_count"),
            "pdf_url": pdf_url,
            "bibtex": bibtex,
            "github_url": cell_value(raw, header_map, "github_url", "github"),
            "seed_bucket": cell_value(raw, header_map, "seed_bucket", "Seed_bucket"),
        }
        target_ws.append([row.get(col, "") for col in GUIDE_COLUMNS])
    style_sheet(target_ws)
    return count


def main() -> None:
    backup = SEED_FILE.with_suffix(f".backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(SEED_FILE, backup)

    src = load_workbook(SEED_FILE, data_only=False)
    dst = Workbook()
    dst.remove(dst.active)

    prefixes = {
        "Survey": "SUR",
        "Image": "IMG",
        "Videos": "VID",
        "Others": "OTH",
    }

    for sheet_name in ["Survey", "Image", "Videos", "Others"]:
        if sheet_name not in src.sheetnames:
            continue
        normalize_sheet(src[sheet_name], dst.create_sheet(sheet_name), prefixes[sheet_name])

    for sheet_name in EXTRA_SHEETS:
        ws = dst.create_sheet(sheet_name)
        ws.append(GUIDE_COLUMNS)
        style_sheet(ws)

    if "Notes" in src.sheetnames:
        ws = dst.create_sheet("Original_Notes")
        for row in src["Notes"].iter_rows(values_only=True):
            ws.append(list(row))

    dst.save(SEED_FILE)
    print(f"Normalized {SEED_FILE}")
    print(f"Backup saved to {backup}")


if __name__ == "__main__":
    main()
