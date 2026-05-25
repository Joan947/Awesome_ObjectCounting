"""Append a supervised Google Scholar seed batch to the normalized seed workbook."""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SEED_FILE = ROOT / "data" / "seed_papers_google_scholar.xlsx"
QUERY_LOG = ROOT / "data" / "query_log.xlsx"


RECORDS = [
    {
        "sheet": "Videos",
        "bucket": "Video",
        "query": '"video object counting"',
        "title": "Efficient masked AutoEncoder for video object counting and a large-scale benchmark",
        "authors": "B Cao; Q Lu; J Feng; Q Wang; P Zhu",
        "year": 2025,
        "venue": "ICLR",
        "url": "https://proceedings.iclr.cc/paper_files/paper/2025/hash/36fa677083cb993e0036b58d2252dcf6-Abstract-Conference.html",
        "pdf": "https://proceedings.iclr.cc/paper_files/paper/2025/file/36fa677083cb993e0036b58d2252dcf6-Paper-Conference.pdf",
        "citation_count": 5,
        "modality": "Video",
        "task_category": "Dataset/Benchmark; Video counting",
        "input_type": "Video",
        "output_type": "Density Map + Count",
        "dataset_or_benchmark": "Video object counting benchmark",
        "main_contribution": "Introduces an efficient masked autoencoder approach and a large-scale benchmark for video object counting.",
    },
    {
        "sheet": "Videos",
        "bucket": "Video",
        "query": '"video object counting"',
        "title": "Open-world object counting in videos",
        "authors": "N Amini-Naieni; A Zisserman",
        "year": 2026,
        "venue": "AAAI",
        "url": "https://ojs.aaai.org/index.php/AAAI/article/view/37214",
        "pdf": "https://ojs.aaai.org/index.php/AAAI/article/view/37214/41176",
        "citation_count": 5,
        "modality": "Video; Multimodal",
        "task_category": "Open-vocabulary; Video tracking-based",
        "input_type": "Video; Text Prompt; Exemplar Image",
        "output_type": "Tracks + Unique Count",
        "dataset_or_benchmark": "",
        "main_contribution": "Defines open-world object counting in videos using text descriptions or image examples to specify target objects.",
    },
    {
        "sheet": "Videos",
        "bucket": "Video",
        "query": '"video object counting"',
        "title": "Video object counting with scene-aware multi-object tracking",
        "authors": "Y Li; L Qu; G Cai; G Cheng; L Qian; Y Dou",
        "year": 2023,
        "venue": "Journal of Database Management",
        "url": "https://www.igi-global.com/article/video-object-counting-with-scene-aware-multi-object-tracking/321553",
        "pdf": "https://www.igi-global.com/viewtitle.aspx?titleid=321553",
        "citation_count": 3,
        "modality": "Video",
        "task_category": "Video tracking-based",
        "input_type": "Video",
        "output_type": "Tracks + Unique Count",
        "dataset_or_benchmark": "",
        "main_contribution": "Uses scene-aware multi-object tracking to count video objects through persistent object IDs.",
    },
    {
        "sheet": "Videos",
        "bucket": "Video",
        "query": '"video object counting"',
        "title": "Video object counting dataset",
        "authors": "OJ Makhura; JC Woods",
        "year": 2019,
        "venue": "IEEE Conference on Multimedia Information Processing and Retrieval",
        "url": "https://ieeexplore.ieee.org/abstract/document/8695318/",
        "pdf": "",
        "citation_count": 3,
        "modality": "Video",
        "task_category": "Dataset/Benchmark",
        "input_type": "Video",
        "output_type": "Benchmark",
        "dataset_or_benchmark": "Video object counting dataset",
        "main_contribution": "Provides a video object counting dataset for evaluating counting methods.",
    },
    {
        "sheet": "Videos",
        "bucket": "Video",
        "query": '"video object counting"',
        "title": "Depth-assisted network for indiscernible marine object counting with adaptive motion-differentiated feature encoding",
        "authors": "C Ma; K Li; S Liu; H Mei",
        "year": 2025,
        "venue": "IEEE Transactions on Circuits and Systems for Video Technology",
        "url": "https://ieeexplore.ieee.org/abstract/document/11121905/",
        "pdf": "https://arxiv.org/pdf/2503.08152",
        "citation_count": 2,
        "modality": "Video; RGB-D",
        "task_category": "Application-specific; Video counting",
        "input_type": "Video; Depth/RGB-D",
        "output_type": "Count Only",
        "dataset_or_benchmark": "Marine object counting",
        "main_contribution": "Counts visually indiscernible marine objects using depth-assisted and motion-differentiated video features.",
    },
    {
        "sheet": "Others",
        "bucket": "3D / Point Cloud",
        "query": '"3D object counting" OR "point cloud object counting"',
        "title": "Countnet3d: A 3d computer vision approach to infer counts of occluded objects",
        "authors": "P Jenkins; K Armstrong; S Nelson",
        "year": 2023,
        "venue": "WACV",
        "url": "https://openaccess.thecvf.com/content/WACV2023/html/Jenkins_CountNet3D_A_3D_Computer_Vision_Approach_To_Infer_Counts_of_WACV_2023_paper.html",
        "pdf": "https://openaccess.thecvf.com/content/WACV2023/papers/Jenkins_CountNet3D_A_3D_Computer_Vision_Approach_To_Infer_Counts_of_WACV_2023_paper.pdf",
        "citation_count": 17,
        "modality": "3D / Point Cloud",
        "task_category": "3D object counting",
        "input_type": "Point Cloud; 3D Data",
        "output_type": "Count Only",
        "dataset_or_benchmark": "",
        "main_contribution": "Uses 3D computer vision to estimate counts of occluded objects with uncertainty.",
    },
    {
        "sheet": "Others",
        "bucket": "3D / Point Cloud",
        "query": '"3D object counting" OR "point cloud object counting"',
        "title": "Stacked objects count based on density map and two vanishing points",
        "authors": "R Yang; Y Zhao; J Shen; X Gong",
        "year": 2022,
        "venue": "SPIE Conference on Information Optics and Photonics",
        "url": "https://www.spiedigitallibrary.org/conference-proceedings-of-spie/12257/122571Y/Stacked-objects-count-based-on-density-map-and-two-vanishing/10.1117/12.2640364.short",
        "pdf": "",
        "citation_count": "",
        "modality": "3D / Point Cloud",
        "task_category": "Density-based; Application-specific",
        "input_type": "Image; 3D Geometry",
        "output_type": "Density Map + Count",
        "dataset_or_benchmark": "Stacked objects",
        "main_contribution": "Counts stacked objects by combining density maps with two-vanishing-point geometry.",
    },
    {
        "sheet": "Remote_Sensing_Aerial_UAV",
        "bucket": "Remote Sensing / Aerial / UAV",
        "query": '"remote sensing" "object counting"',
        "title": "Counting from sky: A large-scale data set for remote sensing object counting and a benchmark method",
        "authors": "G Gao; Q Liu; Y Wang",
        "year": 2020,
        "venue": "IEEE Transactions on Geoscience and Remote Sensing",
        "url": "https://ieeexplore.ieee.org/abstract/document/9200587/",
        "pdf": "https://arxiv.org/pdf/2008.12470",
        "citation_count": 102,
        "modality": "Remote Sensing / Aerial",
        "task_category": "Dataset/Benchmark; Density-based",
        "input_type": "Remote Sensing Image",
        "output_type": "Density Map + Count",
        "dataset_or_benchmark": "RSOC",
        "main_contribution": "Introduces a large-scale remote sensing object counting dataset and benchmark method.",
    },
    {
        "sheet": "Remote_Sensing_Aerial_UAV",
        "bucket": "Remote Sensing / Aerial / UAV",
        "query": '"remote sensing" "object counting"',
        "title": "Remote sensing object counting with online knowledge learning",
        "authors": "S Jiang; Y Gao; B Li; F Cheng",
        "year": 2025,
        "venue": "IEEE Transactions on Geoscience and Remote Sensing",
        "url": "https://ieeexplore.ieee.org/abstract/document/10922124/",
        "pdf": "",
        "citation_count": 5,
        "modality": "Remote Sensing / Aerial",
        "task_category": "Density-based",
        "input_type": "Remote Sensing Image",
        "output_type": "Density Map + Count",
        "dataset_or_benchmark": "",
        "main_contribution": "Applies online knowledge learning and distillation to remote sensing object counting.",
    },
    {
        "sheet": "Remote_Sensing_Aerial_UAV",
        "bucket": "Remote Sensing / Aerial / UAV",
        "query": '"remote sensing" "object counting"',
        "title": "Balanced density regression network for remote sensing object counting",
        "authors": "H Guo; J Gao; Y Yuan",
        "year": 2024,
        "venue": "IEEE Transactions on Geoscience and Remote Sensing",
        "url": "https://ieeexplore.ieee.org/abstract/document/10535416/",
        "pdf": "",
        "citation_count": 24,
        "modality": "Remote Sensing / Aerial",
        "task_category": "Density-based",
        "input_type": "Remote Sensing Image",
        "output_type": "Density Map + Count",
        "dataset_or_benchmark": "",
        "main_contribution": "Uses balanced density regression for object counting in remote sensing scenes.",
    },
    {
        "sheet": "Remote_Sensing_Aerial_UAV",
        "bucket": "Remote Sensing / Aerial / UAV",
        "query": '"remote sensing" "object counting"',
        "title": "A lightweight multiscale feature fusion network for remote sensing object counting",
        "authors": "J Yi; Z Shen; F Chen; Y Zhao; S Xiao",
        "year": 2023,
        "venue": "IEEE Transactions on Geoscience and Remote Sensing",
        "url": "https://ieeexplore.ieee.org/abstract/document/10021616/",
        "pdf": "",
        "citation_count": 59,
        "modality": "Remote Sensing / Aerial",
        "task_category": "Density-based",
        "input_type": "Remote Sensing Image",
        "output_type": "Density Map + Count",
        "dataset_or_benchmark": "",
        "main_contribution": "Introduces a lightweight multiscale feature fusion architecture for remote sensing object counting.",
    },
    {
        "sheet": "Medical_Microscopy_Cell",
        "bucket": "Medical / Microscopy / Cell",
        "query": '"cell counting" "computer vision"',
        "title": "Cell counting by regression using convolutional neural network",
        "authors": "Y Xue; N Ray; J Hugh; G Bigras",
        "year": 2016,
        "venue": "ECCV Workshops",
        "url": "https://link.springer.com/content/pdf/10.1007/978-3-319-46604-0_20.pdf",
        "pdf": "",
        "citation_count": 123,
        "modality": "Medical / Microscopy",
        "task_category": "Regression-based; Application-specific",
        "input_type": "Medical/Microscopy Image",
        "output_type": "Count Only",
        "dataset_or_benchmark": "Cell images",
        "main_contribution": "Uses CNN regression for cell counting in microscopy images.",
    },
    {
        "sheet": "Medical_Microscopy_Cell",
        "bucket": "Medical / Microscopy / Cell",
        "query": '"cell counting" "computer vision"',
        "title": "A method of cell counting based on computer vision",
        "authors": "S Ding; H Zhao; Z Han; Y Tang",
        "year": 2022,
        "venue": "International Conference Proceedings",
        "url": "https://ieeexplore.ieee.org/abstract/document/9907162/",
        "pdf": "",
        "citation_count": 1,
        "modality": "Medical / Microscopy",
        "task_category": "Detection-based; Application-specific",
        "input_type": "Medical/Microscopy Image",
        "output_type": "Boxes + Count",
        "dataset_or_benchmark": "Cell images",
        "main_contribution": "Applies computer vision detection and recognition to automatic cell counting.",
    },
    {
        "sheet": "Medical_Microscopy_Cell",
        "bucket": "Medical / Microscopy / Cell",
        "query": '"cell counting" "computer vision"',
        "title": "A systematic survey on biological cell image segmentation and cell counting techniques in microscopic images using machine learning",
        "authors": "H Singh; H Kaur",
        "year": 2024,
        "venue": "Wireless Personal Communications",
        "url": "https://link.springer.com/article/10.1007/s11277-024-11379-7",
        "pdf": "",
        "citation_count": 10,
        "modality": "Medical / Microscopy",
        "task_category": "Survey",
        "input_type": "Medical/Microscopy Image",
        "output_type": "Count Only; Masks + Count",
        "dataset_or_benchmark": "",
        "main_contribution": "Surveys biological cell segmentation and counting techniques for microscopic images.",
    },
    {
        "sheet": "Medical_Microscopy_Cell",
        "bucket": "Medical / Microscopy / Cell",
        "query": '"cell counting" "computer vision"',
        "title": "Rapid characterization of cell and bacteria counts using computer vision",
        "authors": "F Akkoyun; A Ozcelik",
        "year": 2021,
        "venue": "Turkish Journal of Nature and Science",
        "url": "https://dergipark.org.tr/en/doi/10.46810/tdfd.902441",
        "pdf": "https://dergipark.org.tr/en/download/article-file/1659557",
        "citation_count": 10,
        "modality": "Medical / Microscopy",
        "task_category": "Application-specific",
        "input_type": "Medical/Microscopy Image",
        "output_type": "Count Only",
        "dataset_or_benchmark": "Cell and bacteria images",
        "main_contribution": "Uses computer vision to rapidly characterize cell and bacteria counts.",
    },
    {
        "sheet": "Benchmarks_Datasets",
        "bucket": "Benchmarks / Datasets",
        "query": "FSC-147 OR FSCD-147 OR CountBench OR OmniCount object counting",
        "title": "Open-world text-specified object counting",
        "authors": "N Amini-Naieni; K Amini-Naieni; T Han",
        "year": 2023,
        "venue": "arXiv",
        "url": "https://arxiv.org/abs/2306.01851",
        "pdf": "https://arxiv.org/pdf/2306.01851",
        "citation_count": 83,
        "modality": "Image; Multimodal",
        "task_category": "Open-vocabulary; Dataset/Benchmark",
        "input_type": "Image; Text Prompt",
        "output_type": "Count Only",
        "dataset_or_benchmark": "FSC147-D; FSC-147",
        "main_contribution": "Introduces text-specified open-world counting and FSC147-D descriptions for object counting.",
    },
    {
        "sheet": "Benchmarks_Datasets",
        "bucket": "Benchmarks / Datasets",
        "query": "FSC-147 OR FSCD-147 OR CountBench OR OmniCount object counting",
        "title": "The MixCount Dataset: Bridging the Data Gap for Open-Vocabulary Object Counting",
        "authors": "C Dumery; N Amini-Naieni; S Naini; P Fua",
        "year": 2026,
        "venue": "arXiv",
        "url": "https://arxiv.org/abs/2605.18063",
        "pdf": "https://arxiv.org/pdf/2605.18063",
        "citation_count": "",
        "modality": "Image; Multimodal",
        "task_category": "Dataset/Benchmark; Open-vocabulary",
        "input_type": "Image; Text Prompt",
        "output_type": "Benchmark",
        "dataset_or_benchmark": "MixCount; FSC-147",
        "main_contribution": "Introduces MixCount to reduce the data gap for open-vocabulary object counting.",
    },
    {
        "sheet": "Benchmarks_Datasets",
        "bucket": "Benchmarks / Datasets",
        "query": "FSC-147 OR FSCD-147 OR CountBench OR OmniCount object counting",
        "title": "RealCount: Robust Open-World Object Counting via Duplex Contrastive Learning",
        "authors": "Z Shi; R Liu; J Takahashi; S Jiang",
        "year": 2026,
        "venue": "ICASSP",
        "url": "https://ieeexplore.ieee.org/abstract/document/11462715/",
        "pdf": "",
        "citation_count": "",
        "modality": "Image; Multimodal",
        "task_category": "Open-vocabulary; Few-shot",
        "input_type": "Image; Exemplar Patches",
        "output_type": "Count Only",
        "dataset_or_benchmark": "Robust FSC-147",
        "main_contribution": "Studies robust open-world object counting with external exemplars.",
    },
    {
        "sheet": "Benchmarks_Datasets",
        "bucket": "Benchmarks / Datasets",
        "query": "FSC-147 OR FSCD-147 OR CountBench OR OmniCount object counting",
        "title": "TrueCount: Improving Open-World Object Counting with Visual-Language Models and Dynamic Multi-Modal Inputs",
        "authors": "Z Shi; R Liu; J Takahashi; S Jiang",
        "year": 2025,
        "venue": "ACM Multimedia",
        "url": "https://dl.acm.org/doi/abs/10.1145/3746027.3755426",
        "pdf": "https://dl.acm.org/doi/pdf/10.1145/3746027.3755426",
        "citation_count": 1,
        "modality": "Image; Multimodal",
        "task_category": "Open-vocabulary; Foundation-model",
        "input_type": "Image; Text Prompt; Exemplar Patches",
        "output_type": "Count Only",
        "dataset_or_benchmark": "FSC-147; CountBench",
        "main_contribution": "Improves open-world counting using VLMs and dynamic multimodal inputs.",
    },
    {
        "sheet": "Benchmarks_Datasets",
        "bucket": "Benchmarks / Datasets",
        "query": "FSC-147 OR FSCD-147 OR CountBench OR OmniCount object counting",
        "title": "Zero-shot object counting",
        "authors": "J Xu; H Le; V Nguyen; V Ranjan",
        "year": 2023,
        "venue": "CVPR",
        "url": "http://openaccess.thecvf.com/content/CVPR2023/html/Xu_Zero-Shot_Object_Counting_CVPR_2023_paper.html",
        "pdf": "https://openaccess.thecvf.com/content/CVPR2023/papers/Xu_Zero-Shot_Object_Counting_CVPR_2023_paper.pdf",
        "citation_count": 114,
        "modality": "Image; Multimodal",
        "task_category": "Zero-shot; Open-vocabulary",
        "input_type": "Image; Text Prompt",
        "output_type": "Count Only",
        "dataset_or_benchmark": "FSC-147",
        "main_contribution": "Introduces zero-shot object counting with text-specified categories.",
    },
    {
        "sheet": "Thermal_Event_Camera",
        "bucket": "Thermal / Infrared",
        "query": '"thermal infrared" "crowd counting"',
        "title": "A large-scale drone based thermal infrared benchmark and inception transformer network for crowd counting",
        "authors": "X Wang; T Li; Y Liu; S Yao; Y Liu; N Yang; P Zhu",
        "year": 2025,
        "venue": "Pattern Recognition",
        "url": "https://www.sciencedirect.com/science/article/pii/S0031320325014414",
        "pdf": "",
        "citation_count": 3,
        "modality": "Thermal; Remote Sensing / Aerial",
        "task_category": "Dataset/Benchmark; Density-based",
        "input_type": "Thermal Image; UAV Image",
        "output_type": "Density Map + Count",
        "dataset_or_benchmark": "Drone-based thermal infrared crowd counting dataset",
        "main_contribution": "Introduces a drone-based thermal infrared benchmark and transformer model for crowd counting.",
    },
    {
        "sheet": "Thermal_Event_Camera",
        "bucket": "Thermal / Infrared",
        "query": '"thermal infrared" "crowd counting"',
        "title": "Visual prompt multibranch fusion network for RGB-thermal crowd counting",
        "authors": "B Mu; F Shao; Z Xie; H Chen; Q Jiang",
        "year": 2024,
        "venue": "IEEE Internet of Things Journal",
        "url": "https://ieeexplore.ieee.org/abstract/document/10576691/",
        "pdf": "",
        "citation_count": 30,
        "modality": "Thermal; Multimodal",
        "task_category": "Foundation-model; Density-based",
        "input_type": "Image; Thermal Image",
        "output_type": "Density Map + Count",
        "dataset_or_benchmark": "RGB-thermal crowd counting",
        "main_contribution": "Fuses RGB and thermal inputs with visual prompts for crowd counting.",
    },
    {
        "sheet": "Thermal_Event_Camera",
        "bucket": "Thermal / Infrared",
        "query": '"thermal infrared" "crowd counting"',
        "title": "A new approach for crowd counting and individuals detection using thermal video",
        "authors": "A Hassaan; M Dessouky",
        "year": 2019,
        "venue": "National Radio Science Conference",
        "url": "https://www.researchgate.net/profile/Mohamed-Dessouky/publication/333044613_A_NEW_APPROACH_FOR_CROWD_COUNTING_AND_INDIVIDUALS_DETECTION_USING_THERMAL_VIDEO/links/5cd93a6fa6fdccc9dda6f949/A-NEW-APPROACH-FOR-CROWD-COUNTING-AND-INDIVIDUALS-DETECTION-USING-THERMAL-VIDEO.pdf",
        "pdf": "https://www.researchgate.net/profile/Mohamed-Dessouky/publication/333044613_A_NEW_APPROACH_FOR_CROWD_COUNTING_AND_INDIVIDUALS_DETECTION_USING_THERMAL_VIDEO/links/5cd93a6fa6fdccc9dda6f949/A-NEW-APPROACH-FOR-CROWD-COUNTING-AND-INDIVIDUALS-DETECTION-USING-THERMAL-VIDEO.pdf",
        "citation_count": 1,
        "modality": "Thermal; Video",
        "task_category": "Detection-based; Application-specific",
        "input_type": "Thermal Video",
        "output_type": "Boxes + Count",
        "dataset_or_benchmark": "",
        "main_contribution": "Counts and detects individuals in thermal video.",
    },
    {
        "sheet": "Thermal_Event_Camera",
        "bucket": "Thermal / Infrared",
        "query": '"thermal infrared" "crowd counting"',
        "title": "Object detection in thermal imagery for crowd density estimation",
        "authors": "P Timofeeva",
        "year": 2020,
        "venue": "Thesis",
        "url": "https://www.theseus.fi/handle/10024/337796",
        "pdf": "https://www.theseus.fi/bitstream/handle/10024/337796/Timofeeva_Polina.pdf?sequence=2",
        "citation_count": 1,
        "modality": "Thermal",
        "task_category": "Detection-based; Application-specific",
        "input_type": "Thermal Image",
        "output_type": "Boxes + Count",
        "dataset_or_benchmark": "",
        "main_contribution": "Uses thermal imagery object detection for crowd density estimation.",
    },
    {
        "sheet": "Thermal_Event_Camera",
        "bucket": "Event Camera",
        "query": '"event camera" "crowd counting"',
        "title": "FlyCount: High-speed counting of black soldier flies using neuromorphic sensors",
        "authors": "A James; A Seth; A Marcireau",
        "year": 2024,
        "venue": "IEEE Sensors",
        "url": "https://ieeexplore.ieee.org/abstract/document/10770156/",
        "pdf": "",
        "citation_count": 4,
        "modality": "Event Camera",
        "task_category": "Application-specific",
        "input_type": "Event Camera",
        "output_type": "Count Only",
        "dataset_or_benchmark": "Black soldier fly counting",
        "main_contribution": "Uses neuromorphic/event sensors for high-speed biological-object counting.",
    },
]


def normalize_title(title: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", title.lower()).strip()


def extract_arxiv(*values: str) -> str:
    text = " ".join(v or "" for v in values)
    match = re.search(r"arxiv(?:\.org/abs/|\.org/pdf/|:|/)(\d{4}\.\d{4,5})", text, re.I)
    return match.group(1) if match else ""


def max_existing_id(ws, prefix: str) -> int:
    headers = [cell.value for cell in ws[1]]
    try:
        idx = headers.index("seed_id") + 1
    except ValueError:
        return 0
    max_id = 0
    for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, values_only=True):
        value = str(row[0] or "")
        match = re.search(r"(\d+)$", value)
        if value.startswith(prefix) and match:
            max_id = max(max_id, int(match.group(1)))
    return max_id


def prefix_for_sheet(sheet_name: str) -> str:
    return {
        "Survey": "SUR",
        "Image": "IMG",
        "Videos": "VID",
        "Others": "OTH",
        "Benchmarks_Datasets": "BEN",
        "Remote_Sensing_Aerial_UAV": "REM",
        "Medical_Microscopy_Cell": "MED",
        "Thermal_Event_Camera": "THE",
    }.get(sheet_name, "SEE")


def infer_existing_taxonomy(ws) -> None:
    headers = [cell.value for cell in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}
    if "paper_title" not in col:
        return
    for row_idx in range(2, ws.max_row + 1):
        title = str(ws.cell(row_idx, col["paper_title"]).value or "")
        text = " ".join(
            str(ws.cell(row_idx, col[name]).value or "")
            for name in ["paper_title", "modality", "task_category", "dataset_or_benchmark", "main_contribution", "google_scholar_query"]
            if name in col
        ).lower()
        if not ws.cell(row_idx, col["input_type"]).value:
            if "video" in text:
                ws.cell(row_idx, col["input_type"], "Video")
            elif "point cloud" in text or "3d" in text:
                ws.cell(row_idx, col["input_type"], "Point Cloud; 3D Data")
            elif "rgb-d" in text or "depth" in text:
                ws.cell(row_idx, col["input_type"], "Image; Depth/RGB-D")
            elif "remote sensing" in text or "aerial" in text or "uav" in text:
                ws.cell(row_idx, col["input_type"], "Remote Sensing Image")
            elif "cell" in text or "microscopy" in text:
                ws.cell(row_idx, col["input_type"], "Medical/Microscopy Image")
            elif "text" in text or "open-vocabulary" in text or "zero-shot" in text:
                ws.cell(row_idx, col["input_type"], "Image; Text Prompt")
            elif "few-shot" in text or "exemplar" in text or "class-agnostic" in text or "fsc-147" in text:
                ws.cell(row_idx, col["input_type"], "Image; Exemplar Boxes/Patches")
            else:
                ws.cell(row_idx, col["input_type"], "Image")
        if not ws.cell(row_idx, col["output_type"]).value:
            if "tracking" in text or "track" in text:
                ws.cell(row_idx, col["output_type"], "Tracks + Unique Count")
            elif "density" in text or "crowd" in text:
                ws.cell(row_idx, col["output_type"], "Density Map + Count")
            elif "segmentation" in text or "mask" in text or "sam" in text:
                ws.cell(row_idx, col["output_type"], "Masks + Count")
            elif "benchmark" in text or "dataset" in text:
                ws.cell(row_idx, col["output_type"], "Benchmark")
            else:
                ws.cell(row_idx, col["output_type"], "Count Only")
        if not ws.cell(row_idx, col["main_contribution"]).value and "notes" in col:
            ws.cell(row_idx, col["main_contribution"], ws.cell(row_idx, col["notes"]).value)
        if not ws.cell(row_idx, col["source_found"]).value:
            ws.cell(row_idx, col["source_found"], "Google Scholar")
        if not ws.cell(row_idx, col["notes"]).value:
            ws.cell(row_idx, col["notes"], "Existing manual seed normalized to final guide schema.")
        if "bibtex_collected" in col and not ws.cell(row_idx, col["bibtex_collected"]).value:
            ws.cell(row_idx, col["bibtex_collected"], "Yes" if ws.cell(row_idx, col.get("bibtex", 0)).value else "No")
        if "pdf_saved" in col and not ws.cell(row_idx, col["pdf_saved"]).value:
            ws.cell(row_idx, col["pdf_saved"], "Yes" if ws.cell(row_idx, col.get("pdf_url", 0)).value else "No")


def append_records() -> None:
    wb = load_workbook(SEED_FILE)
    for ws in wb.worksheets:
        infer_existing_taxonomy(ws)

    existing_titles = set()
    for ws in wb.worksheets:
        headers = [cell.value for cell in ws[1]]
        if "paper_title" not in headers:
            continue
        title_col = headers.index("paper_title") + 1
        for row in ws.iter_rows(min_row=2, min_col=title_col, max_col=title_col, values_only=True):
            if row[0]:
                existing_titles.add(normalize_title(str(row[0])))

    appended_by_query: dict[str, int] = {}
    skipped = 0
    for record in RECORDS:
        normalized = normalize_title(record["title"])
        if normalized in existing_titles:
            skipped += 1
            continue
        ws = wb[record["sheet"]]
        headers = [cell.value for cell in ws[1]]
        col = {h: i + 1 for i, h in enumerate(headers)}
        prefix = prefix_for_sheet(record["sheet"])
        next_id = max_existing_id(ws, prefix) + 1
        row = ws.max_row + 1
        values = {
            "seed_id": f"{prefix}{next_id:03d}",
            "paper_title": record["title"],
            "authors": record["authors"],
            "year": record["year"],
            "venue": record["venue"],
            "source_found": "Google Scholar",
            "google_scholar_query": record["query"],
            "scopus_query": "",
            "modality": record["modality"],
            "task_category": record["task_category"],
            "input_type": record["input_type"],
            "output_type": record["output_type"],
            "dataset_or_benchmark": record["dataset_or_benchmark"],
            "main_contribution": record["main_contribution"],
            "relevance_score": "High",
            "doi": "",
            "arxiv_id": extract_arxiv(record["url"], record["pdf"]),
            "paper_url": record["url"],
            "bibtex_collected": "No",
            "pdf_saved": "No",
            "notes": f"Added from visible Google Scholar results on {date.today().isoformat()}.",
            "citation_count": record["citation_count"],
            "pdf_url": record["pdf"],
            "bibtex": "",
            "github_url": "",
            "seed_bucket": record["bucket"],
        }
        for key, value in values.items():
            if key in col:
                ws.cell(row, col[key], value)
        existing_titles.add(normalized)
        appended_by_query[record["query"]] = appended_by_query.get(record["query"], 0) + 1

    qwb = load_workbook(QUERY_LOG)
    qws = qwb.active
    start = qws.max_row + 1
    query_meta = {
        '"video object counting"': ("Video", 10),
        '"3D object counting" OR "point cloud object counting"': ("3D / Point Cloud", 10),
        '"remote sensing" "object counting"': ("Remote Sensing / Aerial / UAV", 10),
        '"cell counting" "computer vision"': ("Medical / Microscopy / Cell", 10),
        "FSC-147 OR FSCD-147 OR CountBench OR OmniCount object counting": ("Benchmarks / Datasets", 10),
        '"thermal infrared" "crowd counting"': ("Thermal / Infrared", 10),
        '"event camera" "crowd counting"': ("Event Camera", 10),
    }
    for offset, (query, saved) in enumerate(appended_by_query.items(), start=0):
        bucket, inspected = query_meta.get(query, ("", 10))
        qws.append(
            [
                f"Q{start + offset:03d}",
                date.today().isoformat(),
                "Google Scholar",
                bucket,
                query,
                "Top visible relevance-sorted results; English scholarly results; no scraping or downloads",
                "relevance",
                inspected,
                saved,
                "",
                "Supervised seed discovery through in-app browser.",
            ]
        )

    wb.save(SEED_FILE)
    qwb.save(QUERY_LOG)
    print(f"Appended {sum(appended_by_query.values())} records; skipped {skipped} duplicates.")
    for query, saved in appended_by_query.items():
        print(f"{saved}: {query}")


if __name__ == "__main__":
    append_records()
