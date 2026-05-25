"""Create a plain modality/application-grouped workbook from literature_matrix_clean.xlsx."""

from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT / "data" / "literature_matrix_clean.xlsx"
OUTPUT_XLSX = ROOT / "outputs" / "literature_matrix_grouped_by_modality.xlsx"

REQUESTED_COLUMNS = [
    "paper_title",
    "authors",
    "year",
    "venue",
    "doi",
    "paper_url",
    "abstract",
    "index_keywords",
    "citation_count",
    "dataset_or_benchmark",
    "modality",
    "task_category",
    "input_type",
    "output_type",
    "relevance_score",
    "tier",
    "github_url",
]

SHEET_ORDER = [
    "Survey",
    "Image",
    "Video",
    "3D_PointCloud",
    "Remote_Sensing_Aerial",
    "Medical_Microscopy_Cell",
    "Thermal_Event",
    "Applications",
    "Multimodal",
    "Other",
]

WIDTHS = {
    "paper_title": 42,
    "authors": 34,
    "year": 10,
    "venue": 28,
    "doi": 22,
    "paper_url": 36,
    "abstract": 70,
    "index_keywords": 46,
    "citation_count": 14,
    "dataset_or_benchmark": 28,
    "modality": 22,
    "task_category": 34,
    "input_type": 24,
    "output_type": 24,
    "relevance_score": 15,
    "tier": 16,
    "github_url": 34,
}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def load_records() -> list[dict[str, object]]:
    wb = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb["included"] if "included" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [clean_text(value) for value in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(value not in (None, "") for value in row):
            continue
        record = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
        records.append(record)
    return records


def has_application_domain(title: str, abstract: str) -> bool:
    title_patterns = [
        r"\bagricultur(?:e|al)\b",
        r"\bagro\b",
        r"\bcrop(?:s)?\b",
        r"\bfruit(?:s)?\b",
        r"\bvegetable(?:s)?\b",
        r"\bflower(?:s)?\b",
        r"\bapple(?:s)?\b",
        r"\borchard(?:s)?\b",
        r"\bwheat\b",
        r"\brice\b",
        r"\bmaize\b",
        r"\bcotton\b",
        r"\btobacco\b",
        r"\bplantation\b",
        r"\bsmart farm\b",
        r"\bfarm(?:ing|s)?\b",
        r"\bpest(?:s)?\b",
        r"\baphid(?:s)?\b",
        r"\bwhitefly\b",
        r"\binsect(?:s)?\b",
        r"\bpoultry\b",
        r"\bchicken(?:s)?\b",
        r"\blivestock\b",
        r"\banimal(?:s)?\b",
        r"\bfish\b",
        r"\bshrimp\b",
        r"\bornamental fish\b",
        r"\baquatic\b",
        r"\bunderwater\b",
        r"\btraffic\b",
        r"\btransportation system(?:s)?\b",
        r"\bvehicle(?:s)?\b",
        r"\bcar object\b",
        r"\bcrowd counting\b",
        r"\bcrowd localization\b",
        r"\bcrowd analysis\b",
        r"\bcrowd density\b",
        r"\bcrowd\b",
        r"\bwarehouse\b",
        r"\binventory\b",
        r"\bstacked goods\b",
        r"\bbuilding material(?:s)?\b",
        r"\bcasting foundr(?:y|ies)\b",
        r"\bsteel sheet(?:s)?\b",
        r"\bwafer\b",
        r"\belectronic part(?:s)?\b",
        r"\belectronic component(?:s)?\b",
        r"\bsealed product(?:s)?\b",
        r"\bindustrial internet\b",
        r"\bindustrial application(?:s)?\b",
        r"\bsmart campus\b",
    ]
    if any(re.search(pattern, title) for pattern in title_patterns):
        return True

    opening_abstract = abstract[:500]
    abstract_patterns = [
        r"\bthis (paper|work|study|research).{0,120}\b(agricultur(?:e|al)|traffic|underwater|warehouse|inventory)\b",
        r"\bwe (propose|present|introduce|develop).{0,120}\b(agricultur(?:e|al)|traffic|underwater|warehouse|inventory)\b",
        r"\b(object counting|counting).{0,80}\bin (agricultur(?:e|al)|traffic|underwater|warehouse|inventory)\b",
    ]
    return any(re.search(pattern, opening_abstract) for pattern in abstract_patterns)


def bucket_for(record: dict[str, object]) -> str:
    modality = clean_text(record.get("modality")).lower()
    task = clean_text(record.get("task_category")).lower()
    title = clean_text(record.get("paper_title")).lower()
    abstract = clean_text(record.get("abstract")).lower()
    dataset = clean_text(record.get("dataset_or_benchmark")).lower()
    input_type = clean_text(record.get("input_type")).lower()
    application = clean_text(record.get("application_area")).lower()
    text = " ".join([modality, task, title, abstract, dataset, input_type, application])
    if "survey" in task or "survey" in title or "review" in title:
        return "Survey"
    if any(term in text for term in ["medical", "microscopy", "cell", "bacteria", "histology"]):
        return "Medical_Microscopy_Cell"
    if any(term in text for term in ["remote", "aerial", "uav", "drone", "satellite"]):
        return "Remote_Sensing_Aerial"
    if any(term in text for term in ["video", "tracking", "temporal"]):
        return "Video"
    if any(term in text for term in ["3d", "point_cloud", "point cloud", "rgb-d", "depth", "lidar"]):
        return "3D_PointCloud"
    if any(term in text for term in ["thermal", "event_camera", "event camera", "infrared"]):
        return "Thermal_Event"
    if has_application_domain(title, abstract):
        return "Applications"
    if any(
        term in text
        for term in [
            "image",
            "single-image",
            "single image",
            "fsc-147",
            "fsc147",
            "few-shot",
            "few shot",
            "class-agnostic",
            "zero-shot",
            "zero shot",
            "open-vocabulary",
            "open vocabulary",
            "text prompt",
            "text-guided",
            "exemplar",
            "density map",
            "object counting",
        ]
    ) or not modality:
        return "Image"
    if "multimodal" in text or "vision-language" in text or "vlm" in text or "clip" in text:
        return "Multimodal"
    return "Other"


def score_tuple(record: dict[str, object]) -> tuple[int, int, str]:
    tier_rank = {"A_core": 0, "B_important": 1, "C_background": 2}
    tier = clean_text(record.get("tier"))
    try:
        citations = int(float(clean_text(record.get("citation_count")).replace(",", "")))
    except Exception:
        citations = 0
    return (tier_rank.get(tier, 9), -citations, clean_text(record.get("paper_title")).lower())


def write_header(ws) -> None:
    for col_idx, column_name in enumerate(REQUESTED_COLUMNS, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = column_name
        cell.alignment = Alignment(vertical="top")


def add_sheet(wb: Workbook, sheet_name: str, records: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(sheet_name)
    write_header(ws)
    records = sorted(records, key=score_tuple)
    for row_idx, record in enumerate(records, start=2):
        for col_idx, column_name in enumerate(REQUESTED_COLUMNS, start=1):
            value = record.get(column_name, "")
            ws.cell(row_idx, col_idx).value = clean_text(value)

    max_row = max(ws.max_row, 1)
    max_col = len(REQUESTED_COLUMNS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.sheet_view.showGridLines = True

    for col_idx, column_name in enumerate(REQUESTED_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = WIDTHS[column_name]
        for cell in ws[letter]:
            cell.alignment = Alignment(vertical="top", wrap_text=column_name in {"abstract", "index_keywords", "task_category", "authors"})
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 42 if row_idx <= 250 else 30


def main() -> None:
    records = load_records()
    grouped = {name: [] for name in SHEET_ORDER}
    for record in records:
        grouped.setdefault(bucket_for(record), []).append(record)

    wb = Workbook()
    del wb[wb.sheetnames[0]]
    for sheet_name in SHEET_ORDER:
        if not grouped.get(sheet_name) and sheet_name in {"Multimodal", "Other"}:
            continue
        add_sheet(wb, sheet_name, grouped.get(sheet_name, []))

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"Created {OUTPUT_XLSX}")
    for sheet_name in SHEET_ORDER:
        print(f"{sheet_name}: {len(grouped.get(sheet_name, []))}")


if __name__ == "__main__":
    main()
