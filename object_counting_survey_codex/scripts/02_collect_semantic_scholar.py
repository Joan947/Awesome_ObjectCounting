"""Enrich the corpus with Semantic Scholar Academic Graph metadata.

Usage:
    S2_API_KEY=... python scripts/02_collect_semantic_scholar.py

The API key is read only from the environment and is never written to disk.
The script respects the approved limit of one request per second.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
RAW_DIR = DATA / "raw_api_results"
MATRIX_XLSX = DATA / "literature_matrix_clean.xlsx"
DEDUP_CSV = RAW_DIR / "deduplicated_unique_records.csv"
QUERY_BANK = ROOT / "query_bank.yaml"
OUT_CSV = RAW_DIR / "semantic_scholar_enriched_records.csv"
SEARCH_CSV = RAW_DIR / "semantic_scholar_search_results.csv"
REF_CSV = RAW_DIR / "semantic_scholar_references.csv"
CIT_CSV = RAW_DIR / "semantic_scholar_citations.csv"
ERROR_CSV = RAW_DIR / "semantic_scholar_errors.csv"
SUMMARY_JSON = RAW_DIR / "semantic_scholar_summary.json"
CACHE_DIR = RAW_DIR / "semantic_scholar_cache"

BASE = "https://api.semanticscholar.org/graph/v1"
RATE_DELAY_SECONDS = 1.1
MAX_BATCH_IDS = 100
MAX_SEARCH_FALLBACK = 120
MAX_LINK_EXPANSION_PAPERS = 80
MAX_LINKS_PER_PAPER = 50

FIELDS = ",".join(
    [
        "paperId",
        "externalIds",
        "url",
        "title",
        "abstract",
        "venue",
        "publicationVenue",
        "year",
        "authors",
        "citationCount",
        "influentialCitationCount",
        "referenceCount",
        "isOpenAccess",
        "citationStyles",
        "fieldsOfStudy",
        "publicationTypes",
        "publicationDate",
        "openAccessPdf",
    ]
)

LINK_FIELDS = (
    "paperId,title,year,venue,publicationVenue,externalIds,url,abstract,"
    "citationCount,influentialCitationCount,authors,fieldsOfStudy,publicationTypes"
)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def norm_doi(value: object) -> str:
    text = clean_text(value)
    text = re.sub(r"^https?://(dx\.)?doi\.org/", "", text, flags=re.I)
    return text.rstrip("}. ").lower()


def norm_arxiv(value: object) -> str:
    text = clean_text(value)
    text = re.sub(r"^arxiv:", "", text, flags=re.I)
    match = re.search(r"(\d{4}\.\d{4,5})(v\d+)?", text)
    return match.group(1) if match else text.lower()


def norm_title(value: object) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def api_key() -> str:
    key = os.environ.get("S2_API_KEY", "").strip()
    if not key:
        raise SystemExit("S2_API_KEY is not set. Export it in the shell before running this script.")
    return key


def request_json(url: str, key: str, *, method: str = "GET", payload: dict | None = None) -> dict | list:
    body = None
    headers = {"x-api-key": key, "User-Agent": "object-counting-survey/0.1"}
    if payload is not None:
        body = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    for attempt in range(5):
        try:
            with urllib.request.urlopen(req, timeout=60) as response:
                time.sleep(RATE_DELAY_SECONDS)
                return json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            if exc.code in {429, 500, 502, 503, 504}:
                wait = max(RATE_DELAY_SECONDS, (2**attempt) * 2)
                time.sleep(wait)
                continue
            raise
        except urllib.error.URLError:
            wait = max(RATE_DELAY_SECONDS, (2**attempt) * 2)
            time.sleep(wait)
    raise RuntimeError(f"Semantic Scholar request failed after retries: {url}")


def load_matrix_records() -> list[dict[str, str]]:
    if not MATRIX_XLSX.exists():
        raise SystemExit(f"Missing {MATRIX_XLSX}")
    wb = load_workbook(MATRIX_XLSX, read_only=True, data_only=True)
    ws = wb["included"] if "included" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [clean_text(h) for h in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(v not in (None, "") for v in row):
            continue
        item = {header[i]: clean_text(row[i] if i < len(row) else "") for i in range(len(header))}
        records.append(item)
    return records


def load_query_bank() -> list[tuple[str, str]]:
    """Parse the simple local query_bank.yaml without adding a YAML dependency."""
    if not QUERY_BANK.exists():
        return []
    queries: list[tuple[str, str]] = []
    category = ""
    for line in QUERY_BANK.read_text(encoding="utf-8").splitlines():
        if not line.strip() or line.lstrip().startswith("#"):
            continue
        if not line.startswith(" ") and line.rstrip().endswith(":"):
            category = line.rstrip()[:-1]
            continue
        match = re.match(r"\s*-\s*(.+?)\s*$", line)
        if match and category:
            queries.append((category, match.group(1).strip().strip("\"'")))
    return queries


def identifier_for(record: dict[str, str]) -> str:
    doi = norm_doi(record.get("doi"))
    arxiv = norm_arxiv(record.get("arxiv_id"))
    if doi:
        return f"DOI:{doi}"
    if arxiv:
        return f"ARXIV:{arxiv}"
    return ""


def chunked(values: list[str], size: int) -> list[list[str]]:
    return [values[i : i + size] for i in range(0, len(values), size)]


def batch_lookup(records: list[dict[str, str]], key: str) -> dict[str, dict]:
    ids = []
    id_to_title = {}
    for record in records:
        ident = identifier_for(record)
        if ident:
            ids.append(ident)
            id_to_title[ident] = norm_title(record.get("paper_title"))
    ids = list(dict.fromkeys(ids))
    results: dict[str, dict] = {}
    for batch in chunked(ids, MAX_BATCH_IDS):
        url = f"{BASE}/paper/batch?fields={urllib.parse.quote(FIELDS)}"
        data = request_json(url, key, method="POST", payload={"ids": batch})
        for ident, paper in zip(batch, data):
            if paper:
                paper["_lookup_id"] = ident
                paper["_source_norm_title"] = id_to_title.get(ident, "")
                results[ident] = paper
        print(f"Semantic Scholar batch: {len(results)}/{len(ids)} identifier records resolved", flush=True)
    return results


def search_fallback(records: list[dict[str, str]], key: str, existing_titles: set[str]) -> list[dict]:
    papers = []
    searched = 0
    for record in records:
        title = clean_text(record.get("paper_title"))
        normalized = norm_title(title)
        if not title or normalized in existing_titles or identifier_for(record):
            continue
        if searched >= MAX_SEARCH_FALLBACK:
            break
        params = urllib.parse.urlencode({"query": title, "limit": 1, "fields": FIELDS})
        url = f"{BASE}/paper/search?{params}"
        try:
            data = request_json(url, key)
        except Exception as exc:
            print(f"Semantic Scholar search failed for title: {title[:80]} ({exc})", flush=True)
            continue
        searched += 1
        for paper in data.get("data", []):
            paper["_lookup_id"] = f"TITLE:{title}"
            paper["_source_norm_title"] = normalized
            papers.append(paper)
            existing_titles.add(normalized)
    print(f"Semantic Scholar search fallback: {len(papers)} papers from {searched} title searches", flush=True)
    return papers


def search_query_bank(key: str, limit: int, errors: list[dict[str, str]]) -> list[dict]:
    papers = []
    seen = set()
    queries = load_query_bank()
    for index, (category, query) in enumerate(queries, start=1):
        params = urllib.parse.urlencode({"query": query, "limit": limit, "fields": FIELDS})
        url = f"{BASE}/paper/search?{params}"
        try:
            data = request_json(url, key)
        except Exception as exc:
            errors.append({"stage": "query_bank_search", "query": query, "error": clean_text(exc)})
            print(f"Semantic Scholar query search failed for {query}: {exc}", flush=True)
            continue
        added = 0
        for paper in data.get("data", []):
            paper_id = clean_text(paper.get("paperId"))
            title_key = norm_title(paper.get("title"))
            dedupe_key = paper_id or title_key
            if not dedupe_key or dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            paper["_lookup_id"] = f"QUERY:{category}:{query}"
            paper["_source_query_category"] = category
            papers.append(paper)
            added += 1
        print(
            f"Semantic Scholar query search: {index}/{len(queries)} queries, "
            f"{added} new records from {query}",
            flush=True,
        )
    return papers


def extract_external_id(external_ids: dict | None, key: str) -> str:
    if not external_ids:
        return ""
    return clean_text(external_ids.get(key) or external_ids.get(key.upper()) or external_ids.get(key.lower()))


def paper_row(paper: dict, source_query: str = "") -> dict[str, str]:
    authors = "; ".join(clean_text(a.get("name")) for a in paper.get("authors", []) if a.get("name"))
    external_ids = paper.get("externalIds") or {}
    open_pdf = paper.get("openAccessPdf") or {}
    publication_venue = paper.get("publicationVenue") or {}
    return {
        "record_id": f"S2_{clean_text(paper.get('paperId'))}",
        "source": "Semantic Scholar",
        "source_query": source_query or clean_text(paper.get("_lookup_id")),
        "paper_title": clean_text(paper.get("title")),
        "authors": authors,
        "year": clean_text(paper.get("year")),
        "venue": clean_text(paper.get("venue") or publication_venue.get("name")),
        "doi": norm_doi(extract_external_id(external_ids, "DOI")),
        "arxiv_id": norm_arxiv(extract_external_id(external_ids, "ArXiv")),
        "paper_url": clean_text(paper.get("url")),
        "abstract": clean_text(paper.get("abstract")),
        "author_keywords": "",
        "index_keywords": "; ".join(paper.get("fieldsOfStudy") or []),
        "citation_count": clean_text(paper.get("citationCount")),
        "influential_citation_count": clean_text(paper.get("influentialCitationCount")),
        "reference_count": clean_text(paper.get("referenceCount")),
        "document_type": "; ".join(paper.get("publicationTypes") or []),
        "dataset_or_benchmark": "",
        "modality": "",
        "task_category": "Semantic Scholar metadata",
        "method_family": "",
        "input_type": "",
        "output_type": "",
        "application_area": "",
        "relevance_score": "",
        "tier": "",
        "reason_kept_or_removed": "Collected from Semantic Scholar API.",
        "semantic_scholar_id": clean_text(paper.get("paperId")),
        "semantic_scholar_url": clean_text(paper.get("url")),
        "open_access_pdf": clean_text(open_pdf.get("url")),
        "is_open_access": clean_text(paper.get("isOpenAccess")),
        "s2_publication_date": clean_text(paper.get("publicationDate")),
        "bibtex": clean_text((paper.get("citationStyles") or {}).get("bibtex")),
    }


def collect_links(
    papers: list[dict],
    key: str,
    relation: str,
    out_path: Path,
    *,
    max_papers: int,
    related_limit: int,
    errors: list[dict[str, str]],
) -> int:
    rows = []
    selected = [p for p in papers if p.get("paperId")][:max_papers]
    for i, paper in enumerate(selected, start=1):
        paper_id = paper["paperId"]
        endpoint = "references" if relation == "reference" else "citations"
        params = urllib.parse.urlencode({"fields": LINK_FIELDS, "limit": related_limit})
        url = f"{BASE}/paper/{urllib.parse.quote(paper_id)}/{endpoint}?{params}"
        try:
            data = request_json(url, key)
        except Exception as exc:
            errors.append({"stage": endpoint, "paper_id": paper_id, "error": clean_text(exc)})
            print(f"Semantic Scholar {endpoint} failed for {paper_id}: {exc}", flush=True)
            continue
        for item in data.get("data", []) or []:
            if not item:
                continue
            linked = item.get("citedPaper") if relation == "reference" else item.get("citingPaper")
            if not linked:
                continue
            row = paper_row(linked, source_query=f"{relation} of {paper_id}")
            row["relation"] = relation
            row["source_paper_id"] = paper_id
            row["source_paper_title"] = clean_text(paper.get("title"))
            rows.append(row)
        print(f"Semantic Scholar {endpoint}: {i}/{len(selected)} seed papers processed", flush=True)
    fields = sorted({k for row in rows for k in row}) or ["relation", "source_paper_id"]
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


def write_rows(path: Path, rows: list[dict[str, str]]) -> None:
    fields = sorted({k for row in rows for k in row}) or ["record_id", "source", "paper_title"]
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--query-limit", type=int, default=20, help="results per query_bank search query")
    parser.add_argument("--max-title-search", type=int, default=MAX_SEARCH_FALLBACK)
    parser.add_argument("--max-expansion-papers", type=int, default=MAX_LINK_EXPANSION_PAPERS)
    parser.add_argument("--related-limit", type=int, default=MAX_LINKS_PER_PAPER)
    parser.add_argument(
        "--skip-links",
        action="store_true",
        help="skip reference/citation expansion after metadata and query search",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    key = api_key()
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    errors: list[dict[str, str]] = []
    records = load_matrix_records()
    batch_results = batch_lookup(records, key)
    resolved = list(batch_results.values())
    resolved_titles = {p.get("_source_norm_title", "") for p in resolved if p.get("_source_norm_title")}
    global MAX_SEARCH_FALLBACK
    MAX_SEARCH_FALLBACK = args.max_title_search
    resolved.extend(search_fallback(records, key, resolved_titles))
    enriched_rows = [paper_row(paper) for paper in resolved if paper.get("title")]
    write_rows(OUT_CSV, enriched_rows)

    search_results = search_query_bank(key, args.query_limit, errors)
    search_rows = [paper_row(paper) for paper in search_results if paper.get("title")]
    write_rows(SEARCH_CSV, search_rows)

    expansion_pool = resolved + search_results
    ref_count = 0
    cit_count = 0
    if not args.skip_links:
        ref_count = collect_links(
            expansion_pool,
            key,
            "reference",
            REF_CSV,
            max_papers=args.max_expansion_papers,
            related_limit=args.related_limit,
            errors=errors,
        )
        cit_count = collect_links(
            expansion_pool,
            key,
            "citation",
            CIT_CSV,
            max_papers=args.max_expansion_papers,
            related_limit=args.related_limit,
            errors=errors,
        )
    else:
        write_rows(REF_CSV, [])
        write_rows(CIT_CSV, [])

    write_rows(ERROR_CSV, errors)
    summary = {
        "input_literature_matrix_records": len(records),
        "metadata_records_written": len(enriched_rows),
        "query_bank_records_written": len(search_rows),
        "references_written": ref_count,
        "citations_written": cit_count,
        "errors_written": len(errors),
        "rate_limit_seconds": RATE_DELAY_SECONDS,
        "query_limit": args.query_limit,
        "max_title_search": args.max_title_search,
        "max_expansion_papers": 0 if args.skip_links else args.max_expansion_papers,
        "related_limit": 0 if args.skip_links else args.related_limit,
        "semantic_scholar_key_source": "S2_API_KEY environment variable",
    }
    SUMMARY_JSON.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2), flush=True)


if __name__ == "__main__":
    main()
