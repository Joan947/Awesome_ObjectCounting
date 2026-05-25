"""Sync enriched seed data into the original Object_Counting.xlsx workbook.

Keeps the original workbook's sheets, columns, and visual style.
"""

from __future__ import annotations

import copy
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
ORIGINAL = ROOT / "Object_Counting.xlsx"
SEED = ROOT / "object_counting_survey_codex" / "data" / "seed_papers_google_scholar.xlsx"


SOURCE_TO_TARGET = {
    "Survey": "Survey",
    "Image": "Image",
    "Videos": "Videos",
    "Others": "Others",
    "Benchmarks_Datasets": "Image",
    "Remote_Sensing_Aerial_UAV": "Others",
    "Medical_Microscopy_Cell": "Others",
    "Thermal_Event_Camera": "Others",
}

FIELD_MAP = {
    "Title": "paper_title",
    "Authors": "authors",
    "Year": "year",
    "Venue": "venue",
    "DOI": "doi",
    "Modality": "modality",
    "URL": "paper_url",
    "Notes": "notes",
    "Seed_bucket": "seed_bucket",
    "Input_type": "input_type",
    "Output_type": "output_type",
    "Dataset": "dataset_or_benchmark",
    "Citation_count": "citation_count",
    "Task_category": "task_category",
    "relevance_score": "relevance_score",
    "pdf_url": "pdf_url",
    "query used": "google_scholar_query",
    "bibtex": "bibtex",
    "github": "github_url",
    "github_url": "github_url",
}


def norm_title(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def headers(ws) -> list[str]:
    return [cell.value for cell in ws[1]]


def row_dict(ws, row_idx: int) -> dict[str, object]:
    head = headers(ws)
    return {name: ws.cell(row_idx, idx + 1).value for idx, name in enumerate(head)}


def seed_rows(seed_wb):
    for source_name, target_name in SOURCE_TO_TARGET.items():
        if source_name not in seed_wb.sheetnames:
            continue
        ws = seed_wb[source_name]
        head = headers(ws)
        if "paper_title" not in head:
            continue
        for idx in range(2, ws.max_row + 1):
            values = row_dict(ws, idx)
            if values.get("paper_title"):
                yield source_name, target_name, values


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.protection:
            dst.protection = copy.copy(src.protection)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.font:
            dst.font = copy.copy(src.font)
        if src.border:
            dst.border = copy.copy(src.border)
    if ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def main() -> None:
    backup = ORIGINAL.with_suffix(f".backup_seed_sync_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(ORIGINAL, backup)

    original_wb = load_workbook(ORIGINAL)
    seed_wb = load_workbook(SEED, data_only=False)

    title_index: dict[str, tuple[str, int]] = {}
    for ws in original_wb.worksheets:
        head = headers(ws) if ws.max_row else []
        if "Title" not in head:
            continue
        title_col = head.index("Title") + 1
        for row_idx in range(2, ws.max_row + 1):
            title = norm_title(ws.cell(row_idx, title_col).value)
            if title:
                title_index[title] = (ws.title, row_idx)

    updated = 0
    appended = 0
    skipped = 0

    for source_name, target_name, seed in seed_rows(seed_wb):
        title_key = norm_title(seed.get("paper_title"))
        if not title_key:
            skipped += 1
            continue

        if title_key in title_index:
            sheet_name, row_idx = title_index[title_key]
            ws = original_wb[sheet_name]
        else:
            ws = original_wb[target_name]
            row_idx = ws.max_row + 1
            style_row = ws.max_row if ws.max_row >= 2 else 1
            copy_row_style(ws, style_row, row_idx)
            title_index[title_key] = (target_name, row_idx)
            appended += 1

        head = headers(ws)
        col = {name: idx + 1 for idx, name in enumerate(head)}
        changed = False
        for original_col, seed_col in FIELD_MAP.items():
            if original_col not in col:
                continue
            value = seed.get(seed_col)
            if value in (None, ""):
                continue
            current = ws.cell(row_idx, col[original_col]).value
            # Preserve existing manual values unless the seed has newer enrichment fields.
            overwrite = original_col in {"DOI", "Notes", "pdf_url", "bibtex", "URL", "Input_type", "Output_type", "Dataset"}
            if current in (None, "") or overwrite:
                ws.cell(row_idx, col[original_col], value)
                changed = True
        if changed and title_key in title_index and title_index[title_key][1] == row_idx:
            updated += 1

    original_wb.save(ORIGINAL)
    print(f"Backup: {backup}")
    print(f"Updated rows: {updated}")
    print(f"Appended rows: {appended}")
    print(f"Skipped rows: {skipped}")


if __name__ == "__main__":
    main()
